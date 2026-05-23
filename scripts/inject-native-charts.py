#!/usr/bin/env python3
"""
Native chart injector for the spreadsheet factory.

Pipeline:
  1. ExcelJS (TypeScript) builds the workbook with cells/formulas/formatting,
     leaves chart anchors empty, writes the chart spec list to JSON.
  2. This script opens the .xlsx via openpyxl, reads the chart spec JSON,
     adds NATIVE chart objects (BarChart / LineChart / DoughnutChart / etc.)
     bound to real cell ranges via openpyxl.chart.Reference, and saves.
  3. Result: charts that auto-update when the buyer edits the underlying
     cells in Excel or Google Sheets.

Args:
    sys.argv[1] = path to the .xlsx file
    sys.argv[2] = path to the chart-spec JSON file

Exit codes:
    0  success
    1  bad input (missing args, malformed JSON, file not found)
    2  openpyxl failure (printed to stderr)

JSON spec shape (one item per chart):
    {
      "tab": "Dashboard",
      "type": "bar"|"column"|"line"|"pie"|"doughnut"|"area",
      "title": "Monthly Income vs Expenses",
      "dataRange": "Dashboard!C18:D29",          # values
      "categoryRange": "Dashboard!B18:B29",      # optional labels
      "anchor": "F5",                            # top-left A1 cell
      "size": { "width": 480, "height": 320 },   # pixels
      "seriesColors": ["5C7558", "E5BAA8"],
      "legend": "b"|"t"|"r"|"l"|"none"
    }
"""

import json
import sys
import re
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.chart import (
        BarChart, LineChart, PieChart, DoughnutChart, AreaChart, Reference,
    )
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.drawing.line import LineProperties
except ImportError as e:
    print(f"ERROR: openpyxl not installed in this Python: {e}", file=sys.stderr)
    sys.exit(2)


CHART_BUILDERS = {
    "bar":      lambda: BarChart(),
    "column":   lambda: BarChart(),       # column = vertical bar
    "line":     lambda: LineChart(),
    "pie":      lambda: PieChart(),
    "doughnut": lambda: DoughnutChart(),
    "area":     lambda: AreaChart(),
}


def parse_a1_range(s, default_sheet=None):
    """
    Parse 'Sheet1!B5:C10' or 'B5:C10' → (sheet, min_col, min_row, max_col, max_row).
    Cell letters are 1-indexed.
    """
    sheet = default_sheet
    if "!" in s:
        sheet_part, range_part = s.split("!", 1)
        sheet = sheet_part.strip().strip("'")
    else:
        range_part = s
    if ":" in range_part:
        tl, br = range_part.split(":", 1)
    else:
        tl = br = range_part
    return (sheet, *col_row(tl), *col_row(br))


def col_row(a1):
    m = re.match(r"^([A-Z]+)(\d+)$", a1.strip(), re.IGNORECASE)
    if not m:
        raise ValueError(f"Bad A1 cell: {a1}")
    letters = m.group(1).upper()
    col = 0
    for ch in letters:
        col = col * 26 + (ord(ch) - 64)
    return col, int(m.group(2))


def hex_to_argb(hex_str):
    h = hex_str.strip().lstrip("#").upper()
    if len(h) == 6:
        return "FF" + h
    if len(h) == 8:
        return h
    return "FF000000"


def apply_series_colors(chart, colors):
    """Best-effort: paint each series in order with the requested hex color."""
    if not colors:
        return
    for i, series in enumerate(chart.series):
        if i >= len(colors):
            break
        argb = hex_to_argb(colors[i])
        # Solid-fill for bars/pie slices/area; line color for lines
        gp = GraphicalProperties(solidFill=argb)
        if isinstance(chart, LineChart):
            gp.line = LineProperties(solidFill=argb, w=20000)
        series.graphicalProperties = gp


def build_chart(spec, ws_default_name):
    ctype = spec.get("type", "column").lower()
    builder = CHART_BUILDERS.get(ctype)
    if not builder:
        raise ValueError(f"Unsupported chart type: {ctype}")
    chart = builder()
    if ctype == "column":
        chart.type = "col"
    elif ctype == "bar":
        chart.type = "bar"

    title = spec.get("title")
    if title:
        chart.title = title

    legend_pos = spec.get("legend")
    if legend_pos == "none":
        chart.legend = None
    elif legend_pos in ("b", "t", "r", "l"):
        # openpyxl uses position codes
        if chart.legend is None:
            from openpyxl.chart.legend import Legend
            chart.legend = Legend()
        chart.legend.position = legend_pos

    # Hide axis titles by default
    if hasattr(chart, "y_axis"):
        chart.y_axis.title = None
    if hasattr(chart, "x_axis"):
        chart.x_axis.title = None

    return chart


def main():
    if len(sys.argv) < 3:
        print("usage: inject-native-charts.py <xlsx> <charts.json>", file=sys.stderr)
        return 1

    xlsx_path = Path(sys.argv[1])
    json_path = Path(sys.argv[2])

    if not xlsx_path.exists():
        print(f"ERROR: xlsx not found: {xlsx_path}", file=sys.stderr)
        return 1
    if not json_path.exists():
        print(f"ERROR: chart spec JSON not found: {json_path}", file=sys.stderr)
        return 1

    try:
        chart_specs = json.loads(json_path.read_text())
    except Exception as e:
        print(f"ERROR: bad chart JSON: {e}", file=sys.stderr)
        return 1

    if not isinstance(chart_specs, list):
        print("ERROR: chart spec JSON must be an array", file=sys.stderr)
        return 1

    wb = load_workbook(xlsx_path)
    added = 0
    skipped = 0

    for spec in chart_specs:
        try:
            tab = spec.get("tab")
            if not tab or tab not in wb.sheetnames:
                print(f"WARN: tab '{tab}' missing; skip", file=sys.stderr)
                skipped += 1
                continue
            ws = wb[tab]

            # Data range
            data_range = spec.get("dataRange")
            if not data_range:
                print(f"WARN: chart missing dataRange on {tab}; skip", file=sys.stderr)
                skipped += 1
                continue
            ds, dmc, dmr, dxc, dxr = parse_a1_range(data_range, default_sheet=tab)
            data_ws = wb[ds] if ds in wb.sheetnames else ws
            data_ref = Reference(data_ws, min_col=dmc, min_row=dmr, max_col=dxc, max_row=dxr)

            cat_range = spec.get("categoryRange")
            cats_ref = None
            if cat_range:
                cs, cmc, cmr, cxc, cxr = parse_a1_range(cat_range, default_sheet=tab)
                cat_ws = wb[cs] if cs in wb.sheetnames else ws
                cats_ref = Reference(cat_ws, min_col=cmc, min_row=cmr, max_col=cxc, max_row=cxr)

            chart = build_chart(spec, ws.title)
            # titles_from_data=True means first row of data is treated as series names
            chart.add_data(data_ref, titles_from_data=True)
            if cats_ref:
                chart.set_categories(cats_ref)

            apply_series_colors(chart, spec.get("seriesColors", []))

            # Size
            sz = spec.get("size", {})
            chart.width = max(8, (sz.get("width", 480) / 64))   # openpyxl chart width in "cm-ish"
            chart.height = max(6, (sz.get("height", 320) / 40))

            anchor = spec.get("anchor", "B2")
            ws.add_chart(chart, anchor)
            added += 1
        except Exception as e:
            print(f"WARN: chart on {spec.get('tab','?')} failed: {e}", file=sys.stderr)
            skipped += 1

    wb.save(xlsx_path)

    # Single-line JSON result on stdout so the TS caller can parse it
    print(json.dumps({"ok": True, "added": added, "skipped": skipped}))
    return 0


if __name__ == "__main__":
    sys.exit(main())
